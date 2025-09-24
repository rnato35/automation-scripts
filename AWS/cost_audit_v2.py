#!/usr/bin/env python3
"""
EC2-Other Cost Audit Script with CSV Export
Generates detailed CSV reports for cost optimization
Supports any AWS region via command line parameter
"""

import argparse
import boto3
import os
from datetime import datetime, timedelta, timezone
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class EC2CostAuditor:
    def __init__(self, region: str, aws_profile: str = None):
        self.region = region
        self.aws_profile = aws_profile
        
        # Create session with profile if specified
        if aws_profile:
            session = boto3.Session(profile_name=aws_profile)
            self.ec2 = session.client('ec2', region_name=region)
            profile_suffix = f"_{aws_profile}"
        else:
            self.ec2 = boto3.client('ec2', region_name=region)
            profile_suffix = ""
            
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.output_filename = f"ec2_audit_{region}{profile_suffix}_{timestamp}.xlsx"
        self.workbook = Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        self.sheets_data = {}
        
    def get_tag_value(self, tags: List[Dict], key: str) -> str:
        """Extract tag value by key, return 'N/A' if not found"""
        if not tags:
            return 'N/A'
        for tag in tags:
            if tag.get('Key') == key:
                return tag.get('Value', 'N/A')
        return 'N/A'
    
    def calculate_ebs_cost(self, size_gb: int, volume_type: str) -> float:
        """Calculate monthly cost for EBS volume based on type and size"""
        cost_per_gb = {
            'gp2': 0.10,
            'gp3': 0.08,
            'io1': 0.125,
            'io2': 0.125,
            'st1': 0.045,
            'sc1': 0.015
        }
        return round(size_gb * cost_per_gb.get(volume_type, 0.045), 2)
    
    def audit_ebs_volumes(self):
        """Audit all EBS volumes and generate CSV reports"""
        print("📊 1. Auditing EBS Volumes...")
        
        response = self.ec2.describe_volumes()
        volumes = response['Volumes']
        
        # All EBS volumes with detailed info
        all_volumes_data = []
        unattached_volumes_data = []
        gp2_migration_data = []
        
        for volume in volumes:
            name = self.get_tag_value(volume.get('Tags', []), 'Name')
            volume_id = volume['VolumeId']
            size = volume['Size']
            volume_type = volume['VolumeType']
            state = volume['State']
            create_time = volume['CreateTime'].strftime('%Y-%m-%d %H:%M:%S')
            encrypted = volume.get('Encrypted', False)
            availability_zone = volume['AvailabilityZone']
            iops = volume.get('Iops', 'N/A')
            throughput = volume.get('Throughput', 'N/A')
            
            # Handle attachments
            attachments = volume.get('Attachments', [])
            instance_id = attachments[0]['InstanceId'] if attachments else 'unattached'
            device = attachments[0]['Device'] if attachments else 'none'
            
            # All volumes data
            all_volumes_data.append([
                volume_id, name, size, volume_type, state, instance_id, device,
                create_time, iops, throughput, encrypted, availability_zone
            ])
            
            # Unattached volumes
            if state == 'available':
                cost = self.calculate_ebs_cost(size, volume_type)
                unattached_volumes_data.append([
                    volume_id, name, size, volume_type, create_time, cost
                ])
            
            # GP2 migration opportunities
            if volume_type == 'gp2':
                gp2_cost = round(size * 0.10, 2)
                gp3_cost = round(size * 0.08, 2)
                savings = round(gp2_cost - gp3_cost, 2)
                gp2_migration_data.append([
                    volume_id, name, size, state, instance_id, gp2_cost, gp3_cost, savings
                ])
        
        # Store data for Excel sheets
        self.sheets_data['EBS_Volumes_All'] = {
            'headers': ['VolumeId', 'Name', 'Size_GB', 'VolumeType', 'State', 'InstanceId', 
                       'Device', 'CreateTime', 'Iops', 'Throughput', 'Encrypted', 'AvailabilityZone'],
            'data': all_volumes_data
        }
        
        self.sheets_data['EBS_Volumes_Unattached'] = {
            'headers': ['VolumeId', 'Name', 'Size_GB', 'VolumeType', 'CreateTime', 'EstimatedMonthlyCost_USD'],
            'data': unattached_volumes_data
        }
        
        self.sheets_data['EBS_GP2_Migration'] = {
            'headers': ['VolumeId', 'Name', 'Size_GB', 'State', 'InstanceId', 
                       'GP2_Monthly_Cost', 'GP3_Monthly_Cost', 'Monthly_Savings'],
            'data': gp2_migration_data
        }
    
    def audit_ebs_snapshots(self):
        """Audit EBS snapshots and generate CSV reports"""
        print("📸 2. Auditing EBS Snapshots...")
        
        response = self.ec2.describe_snapshots(OwnerIds=['self'])
        snapshots = response['Snapshots']
        
        all_snapshots_data = []
        old_snapshots_data = []
        ninety_days_ago = datetime.now(timezone.utc) - timedelta(days=90)
        
        for snapshot in snapshots:
            name = self.get_tag_value(snapshot.get('Tags', []), 'Name')
            snapshot_id = snapshot['SnapshotId']
            volume_id = snapshot.get('VolumeId', 'N/A')
            volume_size = snapshot['VolumeSize']
            start_time = snapshot['StartTime']
            start_time_str = start_time.strftime('%Y-%m-%d %H:%M:%S')
            description = snapshot.get('Description', 'N/A')
            state = snapshot['State']
            progress = snapshot.get('Progress', 'N/A')
            encrypted = snapshot.get('Encrypted', False)
            cost = round(volume_size * 0.05, 2)
            
            # All snapshots data
            all_snapshots_data.append([
                snapshot_id, name, volume_id, volume_size, start_time_str,
                description, state, progress, encrypted, cost
            ])
            
            # Old snapshots (older than 90 days)
            if start_time < ninety_days_ago:
                old_snapshots_data.append([
                    snapshot_id, name, volume_id, volume_size, start_time_str,
                    description, cost
                ])
        
        # Store data for Excel sheets
        self.sheets_data['EBS_Snapshots_All'] = {
            'headers': ['SnapshotId', 'Name', 'VolumeId', 'VolumeSize_GB', 'StartTime',
                       'Description', 'State', 'Progress', 'Encrypted', 'EstimatedMonthlyCost_USD'],
            'data': all_snapshots_data
        }
        
        self.sheets_data['EBS_Snapshots_Old'] = {
            'headers': ['SnapshotId', 'Name', 'VolumeId', 'VolumeSize_GB', 'StartTime',
                       'Description', 'EstimatedMonthlyCost_USD'],
            'data': old_snapshots_data
        }
    
    def audit_elastic_ips(self):
        """Audit Elastic IP addresses and generate CSV reports"""
        print("🌐 3. Auditing Elastic IP Addresses...")
        
        response = self.ec2.describe_addresses()
        addresses = response['Addresses']
        
        all_ips_data = []
        unattached_ips_data = []
        
        for address in addresses:
            name = self.get_tag_value(address.get('Tags', []), 'Name')
            allocation_id = address['AllocationId']
            public_ip = address['PublicIp']
            domain = address['Domain']
            instance_id = address.get('InstanceId', 'none')
            association_id = address.get('AssociationId', 'unattached')
            network_interface_id = address.get('NetworkInterfaceId', 'none')
            status = 'attached' if association_id != 'unattached' else 'UNATTACHED'
            cost = 3.65 if association_id == 'unattached' else 0
            
            # All IPs data
            all_ips_data.append([
                allocation_id, name, public_ip, domain, instance_id,
                association_id, network_interface_id, status, cost
            ])
            
            # Unattached IPs
            if association_id == 'unattached':
                unattached_ips_data.append([
                    allocation_id, name, public_ip, domain, 3.65
                ])
        
        # Store data for Excel sheets
        self.sheets_data['Elastic_IPs_All'] = {
            'headers': ['AllocationId', 'Name', 'PublicIp', 'Domain', 'InstanceId',
                       'AssociationId', 'NetworkInterfaceId', 'Status', 'Monthly_Cost_USD'],
            'data': all_ips_data
        }
        
        self.sheets_data['Elastic_IPs_Unattached'] = {
            'headers': ['AllocationId', 'Name', 'PublicIp', 'Domain', 'Monthly_Waste_USD'],
            'data': unattached_ips_data
        }
    
    def audit_ec2_instances(self):
        """Audit EC2 instances with storage information"""
        print("🖥️  4. Auditing EC2 Instances with Storage...")
        
        response = self.ec2.describe_instances()
        instances_data = []
        
        for reservation in response['Reservations']:
            for instance in reservation['Instances']:
                name = self.get_tag_value(instance.get('Tags', []), 'Name')
                instance_id = instance['InstanceId']
                instance_type = instance['InstanceType']
                state = instance['State']['Name']
                launch_time = instance['LaunchTime'].strftime('%Y-%m-%d %H:%M:%S')
                platform = instance.get('Platform', 'linux')
                
                # Get attached volumes
                volume_ids = []
                for bdm in instance.get('BlockDeviceMappings', []):
                    if 'Ebs' in bdm:
                        volume_ids.append(bdm['Ebs']['VolumeId'])
                
                instances_data.append([
                    instance_id, name, instance_type, state, launch_time,
                    platform, len(volume_ids), ';'.join(volume_ids)
                ])
        
        # Store data for Excel sheet
        self.sheets_data['EC2_Instances'] = {
            'headers': ['InstanceId', 'Name', 'InstanceType', 'State', 'LaunchTime',
                       'Platform', 'AttachedVolumeCount', 'VolumeIds'],
            'data': instances_data
        }
    
    def generate_cost_summary(self):
        """Generate cost summary calculations"""
        print("💰 5. Generating Cost Summary...")
        
        summary_data = [
            ['Category', 'Current_Monthly_Cost_USD', 'Optimization_Opportunity_USD', 'Potential_Savings_Percent'],
            ['Unattached_EBS_Volumes', 'CALCULATED', 'CALCULATED', '100%'],
            ['GP2_to_GP3_Migration', 'CALCULATED', 'CALCULATED', '20%'],
            ['Unattached_Elastic_IPs', 'CALCULATED', 'CALCULATED', '100%'],
            ['Old_Snapshots_90days', 'CALCULATED', 'CALCULATED', '80%'],
            ['Total_EC2_Other_Optimization', 'CALCULATED', 'CALCULATED', '40%']
        ]
        
        # Store data for Excel sheet
        self.sheets_data['Cost_Summary'] = {
            'headers': summary_data[0],
            'data': summary_data[1:]
        }
    
    def generate_recommendations(self):
        """Generate optimization recommendations report"""
        print("📋 6. Generating optimization recommendations...")
        
        recommendations = """# EC2-Other Cost Optimization Recommendations

## Immediate Opportunities (Review Required):

### 1. Unattached EBS Volumes
- Review: ebs_volumes_unattached.csv
- These volumes are not attached to any instances
- Potential Action: Delete after confirming they're not needed
- Risk Level: LOW (verify no important data first)

### 2. Unattached Elastic IPs  
- Review: elastic_ips_unattached.csv
- Each costs $3.65/month when not attached
- Potential Action: Release after confirming not needed
- Risk Level: LOW (can re-allocate new IPs if needed)

### 3. GP2 to GP3 Migration
- Review: ebs_gp2_migration_opportunities.csv
- GP3 offers ~20% cost savings over GP2
- Potential Action: Migrate volumes during maintenance windows
- Risk Level: MEDIUM (requires brief I/O pause during migration)

### 4. Old EBS Snapshots
- Review: ebs_snapshots_old.csv
- Snapshots older than 90 days may not be needed
- Potential Action: Implement automated lifecycle policies
- Risk Level: MEDIUM (verify backup retention requirements)

## Next Steps:
1. Review all CSV files with your team
2. Validate business requirements for each resource
3. Plan migration/cleanup activities during maintenance windows
4. Implement automated policies to prevent future waste

## Important Notes:
- Always test changes in non-production environments first
- Verify business requirements before removing any resources
- Consider implementing AWS Config rules for ongoing governance
"""
        
        # Recommendations are now written in write_excel_file method
        pass
    
    def write_excel_file(self):
        """Write all data to Excel file with multiple sheets"""
        for sheet_name, sheet_data in self.sheets_data.items():
            # Create worksheet
            ws = self.workbook.create_sheet(title=sheet_name)
            
            # Write headers with formatting
            headers = sheet_data['headers']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Write data
            for row, data_row in enumerate(sheet_data['data'], 2):
                for col, value in enumerate(data_row, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        self.workbook.save(self.output_filename)
        
        # Also save recommendations as a separate text file
        recommendations = """# EC2-Other Cost Optimization Recommendations

## Immediate Opportunities (Review Required):

### 1. Unattached EBS Volumes
- Review: EBS_Volumes_Unattached sheet
- These volumes are not attached to any instances
- Potential Action: Delete after confirming they're not needed
- Risk Level: LOW (verify no important data first)

### 2. Unattached Elastic IPs  
- Review: Elastic_IPs_Unattached sheet
- Each costs $3.65/month when not attached
- Potential Action: Release after confirming not needed
- Risk Level: LOW (can re-allocate new IPs if needed)

### 3. GP2 to GP3 Migration
- Review: EBS_GP2_Migration sheet
- GP3 offers ~20% cost savings over GP2
- Potential Action: Migrate volumes during maintenance windows
- Risk Level: MEDIUM (requires brief I/O pause during migration)

### 4. Old EBS Snapshots
- Review: EBS_Snapshots_Old sheet
- Snapshots older than 90 days may not be needed
- Potential Action: Implement automated lifecycle policies
- Risk Level: MEDIUM (verify backup retention requirements)

## Next Steps:
1. Review all Excel sheets with your team
2. Validate business requirements for each resource
3. Plan migration/cleanup activities during maintenance windows
4. Implement automated policies to prevent future waste

## Important Notes:
- Always test changes in non-production environments first
- Verify business requirements before removing any resources
- Consider implementing AWS Config rules for ongoing governance
"""
        
        recommendations_file = self.output_filename.replace('.xlsx', '_recommendations.txt')
        with open(recommendations_file, 'w') as f:
            f.write(recommendations)
    
    def run_audit(self):
        """Run complete EC2 cost audit"""
        profile_info = f" (profile: {self.aws_profile})" if self.aws_profile else " (default profile)"
        print(f"🔍 Starting EC2-Other Cost Audit for {self.region}{profile_info}...")
        print(f"📁 Output file: {self.output_filename}")
        
        self.audit_ebs_volumes()
        self.audit_ebs_snapshots()
        self.audit_elastic_ips()
        self.audit_ec2_instances()
        self.generate_cost_summary()
        self.generate_recommendations()
        
        # Write Excel file
        self.write_excel_file()
        
        print("\n✅ EC2-Other Cost Audit Complete!")
        print(f"📁 Excel report saved to: {self.output_filename}")
        print(f"📋 Recommendations saved to: {self.output_filename.replace('.xlsx', '_recommendations.txt')}")
        
        print(f"\n📊 Excel file contains the following sheets:")
        for sheet_name in self.sheets_data.keys():
            print(f"   • {sheet_name}")
        
        print("\n🎯 Next Steps:")
        print("1. Review ebs_volumes_unattached.csv for immediate savings opportunities")
        print("2. Check elastic_ips_unattached.csv for unused IP addresses")
        print("3. Analyze ebs_gp2_migration_opportunities.csv for GP2→GP3 migration planning")
        print("4. Review ebs_snapshots_old.csv for potential cleanup candidates")
        print("5. Read optimization_recommendations.txt for detailed guidance")
        
        print("\n📈 Key sheets to review first:")
        print("   • EBS_Volumes_Unattached (immediate waste identification)")
        print("   • EBS_GP2_Migration (20% storage savings potential)")
        print("   • Cost_Summary (high-level optimization summary)")
        
        print("\n⚠️  Always validate business requirements before making any changes!")


def main():
    """Main function"""
    parser = argparse.ArgumentParser(
        description='EC2-Other Cost Audit Script with CSV Export',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  python3 cost_audit_v2.py --region us-east-1
  python3 cost_audit_v2.py --region eu-west-1 --aws-profile production
  python3 cost_audit_v2.py --region ap-southeast-1 --aws-profile dev"""
    )
    
    parser.add_argument(
        '--region', 
        required=True,
        help='AWS region to audit (e.g., us-east-1, eu-west-1, ap-southeast-1)'
    )
    
    parser.add_argument(
        '--aws-profile',
        help='AWS profile to use (optional, uses default if not specified)'
    )
    
    args = parser.parse_args()
    
    try:
        # Validate region format
        if not args.region or len(args.region.split('-')) < 3:
            print("❌ Error: Invalid region format. Use format like 'us-east-1'")
            return 1
            
        profile_info = f" with profile '{args.aws_profile}'" if args.aws_profile else " with default profile"
        print(f"🌍 Selected region: {args.region}{profile_info}")
        auditor = EC2CostAuditor(args.region, args.aws_profile)
        auditor.run_audit()
    except Exception as e:
        print(f"❌ Error running audit: {str(e)}")
        print("💡 Make sure you have:")
        print("   • AWS credentials configured (aws configure)")
        if args.aws_profile:
            print(f"   • Valid AWS profile '{args.aws_profile}' configured")
        print("   • Proper permissions for EC2 describe operations")
        print(f"   • Valid region specified: {args.region}")
        return 1
    return 0


if __name__ == "__main__":
    exit(main())